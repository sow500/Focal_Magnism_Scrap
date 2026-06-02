
import io
import os
import re
import json
import queue
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path

import requests
import openpyxl
from flask import Flask, render_template, jsonify, request, Response, send_file

app = Flask(__name__)

EXCEL_PATH = Path(__file__).parent / 'CMT_Paginated_Dataset_Sow.xlsx'
DOWNLOAD_DIR = Path(__file__).parent / 'downloads'
DOWNLOAD_DIR.mkdir(exist_ok=True)

USGS_SEARCH_URL  = 'https://earthquake.usgs.gov/fdsnws/event/1/query'
GLOBALCMT_URL    = 'https://www.globalcmt.org/cgi-bin/globalcmt-cgi-bin/CMT5/form'

# Global state
events = []
status = {}  # {idx: {search, download, event_id, param_url, param_filename, error}}

# SSE clients
sse_clients = []
sse_lock = threading.Lock()

# Prevent concurrent USGS requests
request_semaphore = threading.Semaphore(3)


# ── Startup ─────────────────────────────────────────────────────────────────

def load_events():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        events.append(dict(zip(headers, row)))

    for i in range(len(events)):
        status[i] = {
            'search': 'pending',
            'download': 'pending',
            'event_id': None,
            'param_url': None,
            'param_filename': None,
            'error': None,
            'mt_status': 'pending',
            'mt_data': None,
            'cmt_status': 'pending',
            'cmt_data': None,
        }


# ── Helpers ──────────────────────────────────────────────────────────────────

def parse_event_datetime(date_val, time_str):
    if isinstance(date_val, datetime):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        date_str = str(date_val)[:10]

    time_str = str(time_str).strip()
    if '.' in time_str:
        main, frac = time_str.rsplit('.', 1)
        frac = frac.ljust(6, '0')[:6]
        return datetime.strptime(f'{date_str}T{main}.{frac}', '%Y-%m-%dT%H:%M:%S.%f')
    return datetime.strptime(f'{date_str}T{time_str}', '%Y-%m-%dT%H:%M:%S')


def broadcast(data):
    msg = f"data: {json.dumps(data)}\n\n"
    with sse_lock:
        dead = []
        for q in sse_clients:
            try:
                q.put_nowait(msg)
            except Exception:
                dead.append(q)
        for q in dead:
            sse_clients.remove(q)


# ── USGS logic ────────────────────────────────────────────────────────────────

def search_usgs_event(idx):
    ev = events[idx]
    status[idx].update({'search': 'searching', 'error': None})
    broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})

    with request_semaphore:
        try:
            dt = parse_event_datetime(ev['Date'], ev['Time (UTC)'])
            start_t = (dt - timedelta(minutes=5)).strftime('%Y-%m-%dT%H:%M:%S')
            end_t = (dt + timedelta(minutes=5)).strftime('%Y-%m-%dT%H:%M:%S')
            mw = float(ev['Mw'])

            # Step 1 – find event in catalog
            resp = requests.get(USGS_SEARCH_URL, params={
                'format': 'geojson',
                'starttime': start_t,
                'endtime': end_t,
                'latitude': float(ev['Latitude']),
                'longitude': float(ev['Longitude']),
                'maxradiuskm': 150,
                'minmagnitude': max(0, mw - 0.5),
                'maxmagnitude': mw + 0.5,
                'orderby': 'time',
            }, timeout=30)
            resp.raise_for_status()
            features = resp.json().get('features', [])

            if not features:
                status[idx]['search'] = 'not_found'
                status[idx]['error'] = 'No matching event in USGS catalog'
                broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
                return False

            event_id = features[0]['id']
            status[idx]['event_id'] = event_id

            # Step 2 – get event detail to find finite-fault product
            detail_resp = requests.get(USGS_SEARCH_URL, params={
                'format': 'geojson',
                'eventid': event_id,
            }, timeout=30)
            detail_resp.raise_for_status()
            detail = detail_resp.json()

            products = detail.get('properties', {}).get('products', {})
            ff_list = products.get('finite-fault', [])

            if not ff_list:
                status[idx]['search'] = 'no_finite_fault'
                status[idx]['error'] = 'Event found but has no finite-fault product'
                broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
                return False

            # Step 3 – find .param file in product contents
            contents = ff_list[0].get('contents', {})
            param_entries = {k: v for k, v in contents.items() if k.lower().endswith('.param')}

            if not param_entries:
                all_keys = list(contents.keys())
                status[idx]['search'] = 'no_param'
                status[idx]['error'] = f'finite-fault product found but no .param file. Files: {all_keys[:8]}'
                broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
                return False

            param_key = list(param_entries.keys())[0]
            param_info = param_entries[param_key]
            param_url = param_info.get('url') or param_info.get('downloadUrl')
            safe_name = f"{event_id}_{os.path.basename(param_key).replace('/', '_')}"

            status[idx].update({
                'search': 'found',
                'param_url': param_url,
                'param_filename': safe_name,
            })
            broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
            return True

        except requests.RequestException as e:
            status[idx]['search'] = 'error'
            status[idx]['error'] = f'Network error: {e}'
            broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
            return False
        except Exception as e:
            status[idx]['search'] = 'error'
            status[idx]['error'] = str(e)
            broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
            return False


def download_param_file(idx):
    if status[idx]['search'] != 'found':
        return False

    status[idx]['download'] = 'downloading'
    broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})

    with request_semaphore:
        try:
            param_url = status[idx]['param_url']
            filename = status[idx]['param_filename']
            filepath = DOWNLOAD_DIR / filename

            resp = requests.get(param_url, timeout=60, stream=True)
            resp.raise_for_status()
            with open(filepath, 'wb') as f:
                for chunk in resp.iter_content(chunk_size=8192):
                    f.write(chunk)

            status[idx]['download'] = 'done'
            broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
            return True

        except Exception as e:
            status[idx]['download'] = 'error'
            status[idx]['error'] = str(e)
            broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
            return False


def search_then_download(idx):
    if status[idx]['search'] != 'found':
        search_usgs_event(idx)
    if status[idx]['search'] == 'found' and status[idx]['download'] != 'done':
        download_param_file(idx)


def fetch_moment_tensor(idx):
    # Ensure we have an event_id (search if needed)
    if not status[idx].get('event_id'):
        search_usgs_event(idx)
    if not status[idx].get('event_id'):
        status[idx]['mt_status'] = 'no_event'
        broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
        return False

    status[idx]['mt_status'] = 'fetching'
    broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})

    with request_semaphore:
        try:
            detail_resp = requests.get(USGS_SEARCH_URL, params={
                'format': 'geojson',
                'eventid': status[idx]['event_id'],
            }, timeout=30)
            detail_resp.raise_for_status()
            products = detail_resp.json().get('properties', {}).get('products', {})
            mt_list = products.get('moment-tensor', [])

            if not mt_list:
                status[idx]['mt_status'] = 'not_found'
                status[idx]['error'] = 'No moment-tensor product on USGS'
                broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
                return False

            p = mt_list[0]['properties']
            mw = float(p.get('derived-magnitude') or events[idx]['Mw'])
            # Wells & Coppersmith (1994) scaling relations — all mechanisms
            fault_length = round(10 ** (-2.44 + 0.59 * mw), 1)
            fault_width  = round(10 ** (-1.01 + 0.32 * mw), 1)

            status[idx]['mt_data'] = {
                'latitude':       p.get('derived-latitude'),
                'longitude':      p.get('derived-longitude'),
                'magnitude':      mw,
                'strike1':        p.get('nodal-plane-1-strike'),
                'dip1':           p.get('nodal-plane-1-dip'),
                'rake1':          p.get('nodal-plane-1-rake'),
                'strike2':        p.get('nodal-plane-2-strike'),
                'dip2':           p.get('nodal-plane-2-dip'),
                'rake2':          p.get('nodal-plane-2-rake'),
                'fault_length_km': fault_length,
                'fault_width_km':  fault_width,
                'rupture_direction': 0,
            }
            status[idx]['mt_status'] = 'found'
            broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
            return True

        except Exception as e:
            status[idx]['mt_status'] = 'error'
            status[idx]['error'] = str(e)
            broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
            return False


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html', total=len(events))


@app.route('/api/events')
def get_events():
    page = max(1, int(request.args.get('page', 1)))
    per_page = max(1, int(request.args.get('per_page', 20)))

    start = (page - 1) * per_page
    end = start + per_page
    total = len(events)
    total_pages = max(1, (total + per_page - 1) // per_page)

    page_events = []
    for i, ev in enumerate(events[start:end], start=start):
        date_val = ev['Date']
        date_str = date_val.strftime('%Y-%m-%d') if isinstance(date_val, datetime) else str(date_val)[:10]
        page_events.append({
            'idx': i,
            'date': date_str,
            'time': ev['Time (UTC)'],
            'lat': ev['Latitude'],
            'lon': ev['Longitude'],
            'mw': ev['Mw'],
            'region': ev['Region'],
            'status': status[i],
        })

    return jsonify({
        'events': page_events,
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': total_pages,
    })


@app.route('/api/search/<int:idx>', methods=['POST'])
def search_event(idx):
    if not (0 <= idx < len(events)):
        return jsonify({'error': 'Invalid index'}), 400
    if status[idx]['search'] in ('searching', 'downloading'):
        return jsonify({'status': 'already_running'})
    threading.Thread(target=search_usgs_event, args=(idx,), daemon=True).start()
    return jsonify({'status': 'started'})


@app.route('/api/download/<int:idx>', methods=['POST'])
def download_event(idx):
    if not (0 <= idx < len(events)):
        return jsonify({'error': 'Invalid index'}), 400
    if status[idx].get('download') in ('downloading',):
        return jsonify({'status': 'already_running'})
    threading.Thread(target=search_then_download, args=(idx,), daemon=True).start()
    return jsonify({'status': 'started'})


@app.route('/api/download-all', methods=['POST'])
def download_all():
    def bulk_task():
        indices = list(range(len(events)))
        total_done = sum(1 for i in indices if status[i]['download'] == 'done')
        broadcast({'type': 'bulk_start', 'total': len(indices), 'done': total_done})

        for idx in indices:
            if status[idx]['download'] == 'done':
                continue
            search_then_download(idx)
            time.sleep(0.5)  # gentle rate limiting

        done_count = sum(1 for i in indices if status[i]['download'] == 'done')
        broadcast({'type': 'bulk_done', 'done': done_count, 'total': len(indices)})

    threading.Thread(target=bulk_task, daemon=True).start()
    return jsonify({'status': 'started'})


@app.route('/api/search-all', methods=['POST'])
def search_all():
    def search_task():
        for idx in range(len(events)):
            if status[idx]['search'] in ('found', 'searching'):
                continue
            search_usgs_event(idx)
            time.sleep(0.5)
        broadcast({'type': 'search_all_done'})

    threading.Thread(target=search_task, daemon=True).start()
    return jsonify({'status': 'started'})


@app.route('/api/status-summary')
def status_summary():
    counts = {
        'total': len(events),
        'search_found': sum(1 for s in status.values() if s['search'] == 'found'),
        'search_not_found': sum(1 for s in status.values() if s['search'] in ('not_found', 'no_finite_fault', 'no_param')),
        'search_error': sum(1 for s in status.values() if s['search'] == 'error'),
        'download_done': sum(1 for s in status.values() if s['download'] == 'done'),
        'download_error': sum(1 for s in status.values() if s['download'] == 'error'),
    }
    return jsonify(counts)


@app.route('/api/stream')
def stream():
    def generate():
        q = queue.Queue(maxsize=200)
        with sse_lock:
            sse_clients.append(q)
        # Send current state snapshot
        yield f"data: {json.dumps({'type': 'init', 'status': status})}\n\n"
        try:
            while True:
                try:
                    msg = q.get(timeout=25)
                    yield msg
                except queue.Empty:
                    yield 'data: {"type":"ping"}\n\n'
        finally:
            with sse_lock:
                if q in sse_clients:
                    sse_clients.remove(q)

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


@app.route('/api/get-file/<path:filename>')
def get_file(filename):
    filepath = DOWNLOAD_DIR / filename
    if filepath.exists() and filepath.is_file():
        return send_file(str(filepath.resolve()), as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


@app.route('/api/reset/<int:idx>', methods=['POST'])
def reset_event(idx):
    if not (0 <= idx < len(events)):
        return jsonify({'error': 'Invalid index'}), 400
    status[idx].update({
        'search': 'pending',
        'download': 'pending',
        'event_id': None,
        'param_url': None,
        'param_filename': None,
        'error': None,
        'mt_status': 'pending',
        'mt_data': None,
        'cmt_status': 'pending',
        'cmt_data': None,
    })
    broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
    return jsonify({'status': 'reset'})


def fetch_cmt(idx):
    ev = events[idx]
    status[idx]['cmt_status'] = 'fetching'
    broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})

    with request_semaphore:
        try:
            dt  = parse_event_datetime(ev['Date'], ev['Time (UTC)'])
            mw  = float(ev['Mw'])
            lat = float(ev['Latitude'])
            lon = float(ev['Longitude'])

            resp = requests.get(GLOBALCMT_URL, params={
                'itype': 'ymd',
                'yr': dt.year, 'mo': dt.month, 'day': dt.day,
                'otype': 'nd', 'nday': 2,
                'lmw': max(0, mw - 0.5), 'umw': mw + 0.5,
                'llat': lat - 2, 'ulat': lat + 2,
                'llon': lon - 2, 'ulon': lon + 2,
                'list': 0,
            }, timeout=30)
            resp.raise_for_status()
            html = resp.text

            if 'Fault plane' not in html:
                status[idx]['cmt_status'] = 'not_found'
                status[idx]['error'] = 'No matching event in GlobalCMT catalog'
                broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
                return False

            lat_m  = re.search(r'Lat=\s*(-?\d+\.?\d*)', html)
            lon_m  = re.search(r'Lon=\s*(-?\d+\.?\d*)', html)
            mw_m   = re.search(r'Mw\s*=\s*(\d+\.?\d*)', html)
            planes = re.findall(r'Fault plane:\s+strike=(\d+)\s+dip=(\d+)\s+slip=(-?\d+)', html)

            if not lat_m or len(planes) < 1:
                status[idx]['cmt_status'] = 'not_found'
                status[idx]['error'] = 'Could not parse GlobalCMT response'
                broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
                return False

            mw_val = float(mw_m.group(1)) if mw_m else mw
            fault_length = round(10 ** (-2.44 + 0.59 * mw_val), 1)
            fault_width  = round(10 ** (-1.01 + 0.32 * mw_val), 1)

            status[idx]['cmt_data'] = {
                'latitude':        float(lat_m.group(1)),
                'longitude':       float(lon_m.group(1)) if lon_m else None,
                'magnitude':       mw_val,
                'strike1':         int(planes[0][0]),
                'dip1':            int(planes[0][1]),
                'rake1':           int(planes[0][2]),
                'strike2':         int(planes[1][0]) if len(planes) > 1 else None,
                'dip2':            int(planes[1][1]) if len(planes) > 1 else None,
                'rake2':           int(planes[1][2]) if len(planes) > 1 else None,
                'fault_length_km': fault_length,
                'fault_width_km':  fault_width,
                'rupture_direction': 0,
            }
            status[idx]['cmt_status'] = 'found'
            broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
            return True

        except Exception as e:
            status[idx]['cmt_status'] = 'error'
            status[idx]['error'] = str(e)
            broadcast({'type': 'status', 'idx': idx, 'status': status[idx]})
            return False


@app.route('/api/fetch-mt/<int:idx>', methods=['POST'])
def fetch_mt(idx):
    if not (0 <= idx < len(events)):
        return jsonify({'error': 'Invalid index'}), 400
    if status[idx].get('mt_status') == 'fetching':
        return jsonify({'status': 'already_running'})
    threading.Thread(target=fetch_moment_tensor, args=(idx,), daemon=True).start()
    return jsonify({'status': 'started'})


@app.route('/api/fetch-mt-all', methods=['POST'])
def fetch_mt_all():
    def task():
        for idx in range(len(events)):
            if status[idx].get('mt_status') == 'found':
                continue
            fetch_moment_tensor(idx)
            time.sleep(0.5)
        broadcast({'type': 'mt_all_done'})
    threading.Thread(target=task, daemon=True).start()
    return jsonify({'status': 'started'})


@app.route('/api/fetch-cmt/<int:idx>', methods=['POST'])
def fetch_cmt_route(idx):
    if not (0 <= idx < len(events)):
        return jsonify({'error': 'Invalid index'}), 400
    if status[idx].get('cmt_status') == 'fetching':
        return jsonify({'status': 'already_running'})
    threading.Thread(target=fetch_cmt, args=(idx,), daemon=True).start()
    return jsonify({'status': 'started'})


@app.route('/api/fetch-cmt-all', methods=['POST'])
def fetch_cmt_all():
    def task():
        for idx in range(len(events)):
            if status[idx].get('cmt_status') == 'found':
                continue
            fetch_cmt(idx)
            time.sleep(0.5)
        broadcast({'type': 'cmt_all_done'})
    threading.Thread(target=task, daemon=True).start()
    return jsonify({'status': 'started'})


@app.route('/api/export-mt')
def export_mt():
    HEADERS = [
        '#', 'Date', 'Time (UTC)', 'Region', 'USGS Event ID',
        'Epicenter Lat', 'Epicenter Lon', 'Mw',
        'Strike NP1', 'Dip NP1', 'Rake NP1',
        'Strike NP2', 'Dip NP2', 'Rake NP2',
        'Fault Length (km)', 'Fault Width (km)', 'Rupture Direction',
    ]

    wb = openpyxl.Workbook()
    ws_mt  = wb.active;  ws_mt.title  = 'USGS Moment Tensor'
    ws_cmt = wb.create_sheet('GlobalCMT')
    ws_mt.append(HEADERS)
    ws_cmt.append(HEADERS)

    for i, ev in enumerate(events):
        date_val = ev['Date']
        date_str = date_val.strftime('%Y-%m-%d') if isinstance(date_val, datetime) else str(date_val)[:10]
        base = [i + 1, date_str, str(ev['Time (UTC)']), ev['Region'], status[i].get('event_id', '')]

        mt = status[i].get('mt_data')
        if mt:
            ws_mt.append(base + [
                mt['latitude'], mt['longitude'], mt['magnitude'],
                mt['strike1'], mt['dip1'], mt['rake1'],
                mt['strike2'], mt['dip2'], mt['rake2'],
                mt['fault_length_km'], mt['fault_width_km'], mt['rupture_direction'],
            ])

        cmt = status[i].get('cmt_data')
        if cmt:
            ws_cmt.append(base + [
                cmt['latitude'], cmt['longitude'], cmt['magnitude'],
                cmt['strike1'], cmt['dip1'], cmt['rake1'],
                cmt['strike2'], cmt['dip2'], cmt['rake2'],
                cmt['fault_length_km'], cmt['fault_width_km'], cmt['rupture_direction'],
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=focal_mechanism_data.xlsx'},
    )


# ── Main ──────────────────────────────────────────────────────────────────────

load_events()

if __name__ == '__main__':
    print(f"Loaded {len(events)} earthquake events from Excel.")
    print("Open http://localhost:5001 in your browser.")
    app.run(debug=False, port=5001, threaded=True)
